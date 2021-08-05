import pyodbc
import openpyxl
import unittest
import HtmlTestRunner


class TestL(unittest.TestCase):


        def test123(self):


            wb = openpyxl.load_workbook('D:/OrthoASC/source/OrthoASC_0180_InsertPreadmission.xlsx')
            ws = wb.worksheets[0]
            cols = ws.max_column
            specialty = "Orthopedic ASC"

            thisList = []
            for z in range(1, cols + 1):
                ele = ws.cell(1, z).value
                thisList.insert(z, ele)
            newcols = cols + cols
            x = 0
            for y in range(cols + 1, newcols + 1):
                ws.cell(1, y).value = thisList[x]
                x += 1
            b = 0
            for status in range(newcols + 1, newcols + cols + 1):
                ws.cell(1, status).value = thisList[b]
                b += 1

                # specialty = "Orthopedic ASC"
                query = """SELECT m.procedure_name, i.category, i.description
FROM ngkbm_proc_activity_items_ i
	JOIN ngkbm_proc_activity_xref_ x ON i.practice_id = x.practice_id AND i.instruction_id = x.activity_id
	JOIN ngkbm_proc_mstr_ m ON x.practice_id = m.practice_id AND x.procedure_id = m.procedure_id
 WHERE i.practice_id = '0001' AND specialty = 'Orthopedic ASC' AND activity_type = 'Patient Instructions'
 ORDER BY 1, 2, 3"""

                conn = pyodbc.connect(driver='{SQL Server Native Client 11.0}',
                                      server='NGBLR-03070\MSSQL2014',
                                      database='ACE_Ortho',
                                      uid='sa', pwd='nextgen')
                cursor = conn.cursor()
                cursor.execute(query)

                onerow = cursor.fetchall()
                for x in range(0, len(onerow)):
                    procedure_name = onerow[x].procedure_name
                    category = onerow[x].category
                    description = onerow[x].description

                    ws.cell(row=x + 2, column=cols + 1).value = procedure_name
                    ws.cell(row=x + 2, column=cols + 2).value = category
                    ws.cell(row=x + 2, column=cols + 3).value = description
                wb.save('D:/OrthoASC/Results/Results/OrthoASC_02_DocLog_output_log_last_patientInstructions.xlsx')
                colsnew = ws.max_column
                for s in range(2, ws.max_row + 1):
                    for s2 in range(1, cols + 1):
                        if str(ws.cell(s, s2).value) == str(ws.cell(s, cols + s2).value):
                            ws.cell(s, cols + cols + s2).value = "Pass"
                        elif (ws.cell(s, s2).value is None) and (ws.cell(s, cols + s2).value == ""):
                            ws.cell(s, cols + cols + s2).value = "Pass"
                            print("pass", s, s2, ws.cell(s, s2).value)
                        else:
                            ws.cell(s, cols + cols + s2).value = "Fail"
                            print(s, s2, cols + s2, "Expected :", ws.cell(s, s2).value, "Actual :",
                                  ws.cell(s, cols + s2).value)

                wb.save('D:/OrthoASC/Results/Results/OrthoASC_02_DocLog_output_log_last_patientInstructions.xlsx')


if __name__ == '__main__':
    unittest.main(
        testRunner=HtmlTestRunner.HTMLTestRunner(output='D:/Back up/Back up/pyprojects/PATIENTPORTAL/Reports'))
    # driver.quit()

