import pyodbc
import openpyxl
import unittest
import HtmlTestRunner


class TestL(unittest.TestCase):




    # wb=openpyxl.load_workbook('D:/OrthoASC/source/OrthoASC_04_InsertProcedureOrders.xlsx')
    # ws=wb.active
    # cols=ws.max_column
    # ws.cell(1,cols+1).value="procedure_name"
    # ws.cell(1,cols+2).value="actual_sort_order"
    # ws.cell(1,cols+3).value="actual_class"
    # ws.cell(1,cols+4).value="actual_activity_type"
    # ws.cell(1,cols+5).value="actual_ctegory"
    # ws.cell(1,cols+6).value="actualdesc"
    # ws.cell(1,cols+7).value="procname_status"
    # ws.cell(1,cols+8).value="sortorder_status"
    # ws.cell(1,cols+9).value="class_status"
    # ws.cell(1,cols+10).value="activity_type_status"
    # ws.cell(1,cols+11).value="catergory_status"
    # ws.cell(1,cols+12).value="desc_status"
    #specialty="Orthopedic ASC"

        def test_procedureorder(self):
            wb = openpyxl.load_workbook('D:/OrthoASC/source/OrthoASC_0110_InsertProcedure.xlsx')
            ws = wb.active
            cols = ws.max_column

            thisList = []
            for z in range(1, cols + 1):
                ele = ws.cell(1, z).value
                thisList.insert(z, ele)
            newcols = cols + cols
            x = 0
            for y in range(cols + 1, newcols + 1):
                ws.cell(1, y).value = thisList[x]
                x += 1
            #
            #wb.save('D:/OrthoASC/source/OrthoASC_04_InsertProcedureOrderstest12.xlsx')
            specialty = "Orthopedic ASC"
            query="""SELECT spec.specialty,procedure_name,spec.procedure_id,procedure_site, phx.hpi_reason FROM ngkbm_proc_mstr_ mstr JOIN ngkbm_proc_specialty_xref_ spec on mstr.procedure_id = spec.procedure_id and mstr.practice_id = spec.practice_id JOIN ngkbm_proc_site_xref_ site on mstr.practice_id = site.practice_id and spec.site_id = site.site_id JOIN ngkbm_proc_hpi_xref_ phx ON phx.practice_id = site.practice_id AND phx.proc_site_id = site.site_id AND spec.specialty='Orthopedic ASC' AND spec.practice_id='0001' AND mstr.kbm_ind='N' order by 2,4,5"""
            conn = pyodbc.connect(driver='{SQL Server Native Client 11.0}',
                                               server='NGBLR-03070\MSSQL2014',
                                               database='ACE_Ortho',
                                               uid='sa',pwd='nextgen')
            cursor = conn.cursor()
            cursor.execute(query)
            onerow =cursor.fetchall()
            for x in range(0,len(onerow)):
                data = onerow[x]

           # print("oneline",oneline)
            wb.save('D:/OrthoASC/Results/Results/OrthoASC_04_InsertProcedure_Output.xlsx')


            for x in range(0,len(onerow)):
                target_specialty=onerow[x].specialty
                targetvalue_proc_name=onerow[x].procedure_name
                #targetvalueclass=onerow[x].class
                target_procedure_id = onerow[x].procedure_id
                target_procedure_site = onerow[x].procedure_site
                target_hpi_reason = onerow[x].hpi_reason
                #targetdesc = onerow[x].description
                ws.cell(row=x+2, column=cols+1).value = target_specialty
                ws.cell(row=x+2, column=cols+2).value=targetvalue_proc_name
                ws.cell(row=x+2, column=cols+3).value = target_procedure_id
                ws.cell(row=x+2, column=cols+4).value = target_procedure_site
                ws.cell(row=x+2, column=cols+5).value = target_hpi_reason


                print(onerow[x])
                #print(ws.max_row)
                print(cols)
            wb.save('D:/OrthoASC/Results/Results/OrthoASC_04_InsertProcedure_Output.xlsx')
            for s in range(2,ws.max_row+1):

                if ws.cell(s,1).value==ws.cell(s,cols+1).value:
                    ws.cell(s,cols+cols+1).value="Pass"
                else:
                    ws.cell(s, cols+cols+1).value ="Fail"
                if ws.cell(s,2).value==ws.cell(s,cols+2).value:
                    ws.cell(s,cols+cols+2).value="Pass"
                else:
                    ws.cell(s, cols+cols+2).value ="Fail"
                if ws.cell(s, 3).value == ws.cell(s, cols+3).value:
                    ws.cell(s, cols+cols+3).value = "Pass"
                else:
                    ws.cell(s, cols+cols+3).value = "Fail"
                if ws.cell(s, 4).value == ws.cell(s, cols+4).value:
                    ws.cell(s, cols+cols+4).value = "Pass"
                else:
                    ws.cell(s, cols+cols+4).value = "Fail"
                if ws.cell(s, 5).value == ws.cell(s, cols+5).value:
                    ws.cell(s, cols+cols+5).value = "Pass"
                else:
                    ws.cell(s, cols+cols+5).value = "Fail"
            wb.save('D:/OrthoASC/Results/Results/OrthoASC_04_InsertProcedure_Output.xlsx')

        if __name__== '__main__':
            unittest.main(testRunner=HtmlTestRunner.HTMLTestRunner(output='..//Reports'))
         #driver.quit()

