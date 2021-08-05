import pyodbc
import openpyxl

wb=openpyxl.load_workbook('D:/OrthoASC/Results/Results/OrthoASC_02_InsertIndicationsz_output.xlsx')
ws=wb._sheets[0]

print(ws.cell(3,16).value)
print(ws.cell(3,35).value)

