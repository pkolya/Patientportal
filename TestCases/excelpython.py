import pyodbc
import openpyxl

wb=openpyxl.load_workbook('D:/automation/ambulatory.xlsx')
ws=wb.active
ws.cell(row=1,column=5).value="Actual1"
print(ws.cell(row=1,column=4).value)
wb.save('D:/automation/ambulatoryResults.xlsx')