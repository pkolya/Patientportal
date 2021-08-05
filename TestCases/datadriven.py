import pyodbc
import openpyxl

wb=openpyxl.load_workbook('D:/automation/ambulatory.xlsx')
ws=wb.active
ws.cell(1,5).value="Actual_timeout_element"
ws.cell(1,6).value="status"
query="select specialty,panel_name,sort_order,timeout_element from timeout_ext_ where specialty ='Ambulatory Surgery' order by 2"
conn = pyodbc.connect(driver='{SQL Server Native Client 11.0}',
                                   server='NGBLR-03070\MSSQL2014',
                                   database='CG_Merge',
                                   uid='sa',pwd='nextgen')
cursor = conn.cursor()
cursor.execute(query)

onerow=cursor.fetchall()
for x in range(0,len(onerow)):
    targetvaluetimeout=onerow[x].timeout_element
    #taegetvaluepanelname=onerow[x].panel_name
    ws.cell(row=x+2, column=5).value=targetvaluetimeout
wb.save('D:/automation/ambulatorylast.xlsx')
for s in range(2,ws.max_row+1):
    if ws.cell(s,4).value==ws.cell(s,5).value:
        ws.cell(s,6).value="Pass"
    else:
        ws.cell(s, 6).value ="Fail"
wb.save('D:/automation/ambulatorylastresultdemo.xlsx')



