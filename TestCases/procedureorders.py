import pyodbc
import openpyxl
import unittest
import HtmlTestRunner

class TestL(unittest.TestCase):


        # wb=openpyxl.load_workbook('D:/OrthoASC/source/OrthoASC_04_InsertProcedureOrders.xlsx')
        # ws=wb.active
        # cols=ws.max_column
        # ws.cell(1,cols+1).value="procedure_name"
        # ws.cell(1,cols+2).value="actual_sort_order"
        # ws.cell(1,cols+3).value="actual_class"
        # ws.cell(1,cols+4).value="actual_activity_type"
        # ws.cell(1,cols+5).value="actual_ctegory"
        # ws.cell(1,cols+6).value="actualdesc"
        # ws.cell(1,cols+7).value="procname_status"
        # ws.cell(1,cols+8).value="sortorder_status"
        # ws.cell(1,cols+9).value="class_status"
        # ws.cell(1,cols+10).value="activity_type_status"
        # ws.cell(1,cols+11).value="catergory_status"
        # ws.cell(1,cols+12).value="desc_status"
        #specialty="Orthopedic ASC"

        def test_procedureorder(self):
            wb = openpyxl.load_workbook('D:/OrthoASC/source/OrthoASC_04_InsertProcedureOrders.xlsx')
            ws = wb.active
            cols = ws.max_column
            ws.cell(1, cols+1).value = "procedure_name"
            ws.cell(1, cols+2).value = "actual_sort_order"
            ws.cell(1, cols+3).value = "actual_class"
            ws.cell(1, cols+4).value = "actual_activity_type"
            ws.cell(1, cols+5).value = "actual_ctegory"
            ws.cell(1, cols+6).value = "actualdesc"
            ws.cell(1, cols+7).value = "procname_status"
            ws.cell(1, cols+8).value = "sortorder_status"
            ws.cell(1, cols+9).value = "class_status"
            ws.cell(1, cols+10).value = "activity_type_status"
            ws.cell(1, cols+11).value = "catergory_status"
            ws.cell(1, cols+12).value = "desc_status"
            specialty = "Orthopedic ASC"
            query=" select procedure_name,sort_order,class,activity_type,category,description FROM ngkbm_proc_activity_items_ AS npa INNER JOIN ngkbm_proc_mstr_ npm ON npa.practice_id = npm.practice_id AND npm.procedure_id = npa.proc_id where npa.specialty='"+specialty+"' AND npa.practice_id='0001' AND class in ('General Procedure Order','Medication Procedure Order') order by 1,2"
            conn = pyodbc.connect(driver='{SQL Server Native Client 11.0}',
                                               server='NGBLR-03070\MSSQL2014',
                                               database='ACE_Ortho',
                                               uid='sa',pwd='nextgen')
            cursor = conn.cursor()
            cursor.execute(query)

            onerow=cursor.fetchall()
            for x in range(0,len(onerow)):
                targetvalueprocname=onerow[x].procedure_name
                targetvaluelower=targetvalueprocname.lower()
                #targetvalueclass=onerow[x].class
                targetsortorder = onerow[x].sort_order
                targetacttype = onerow[x].activity_type
                targetcategory = onerow[x].category
                targetdesc = onerow[x].description
                ws.cell(row=x+2, column=cols+1).value=targetvaluelower
                ws.cell(row=x+2, column=cols+2).value = targetsortorder
                # ws.cell(row=x+2, column=cols+3).value = onerow[x].class
                ws.cell(row=x+2, column=cols+4).value = targetacttype
                ws.cell(row=x+2, column=cols+5).value = targetcategory
                ws.cell(row=x+2, column=cols+6).value = "'"+targetdesc

                print(ws.cell(row=x+2, column=cols+6).value )
                #print(ws.max_row)
                print(cols)
            wb.save('D:/OrthoASC/Results/OrthoASC_04_InsertProcedureOrdersupdated.xlsx')
            for s in range(2,ws.max_row+1):

                if ws.cell(s,1).value==ws.cell(s,cols+1).value:
                    ws.cell(s,cols+7).value="Pass"
                else:
                    ws.cell(s, cols+7).value ="Fail"
                if ws.cell(s,2).value==ws.cell(s,cols+2).value:
                    ws.cell(s,cols+8).value="Pass"
                else:
                    ws.cell(s, cols+8).value ="Fail"
                if ws.cell(s, 4).value == ws.cell(s, cols+4).value:
                    ws.cell(s, cols+10).value = "Pass"
                else:
                    ws.cell(s, cols+10).value = "Fail"
                if ws.cell(s, 5).value == ws.cell(s, cols+5).value:
                    ws.cell(s, cols+11).value = "Pass"
                else:
                    ws.cell(s, cols+11).value = "Fail"
                abc = ""
                defi = ""
                ws.cell(s, 6).value = abc
                ws.cell(s, cols + 6).value = defi
                definew = defi[1:]
                if abc == definew:
                    ws.cell(s, cols+12).value = "Pass"
                else:
                    ws.cell(s, cols+12).value = "Fail"
            wb.save('D:/OrthoASC/Results/OrthoASC_04_InsertProcedureOrdersResultsordered2.xlsx')

        if __name__== '__main__':
            unittest.main(testRunner=HtmlTestRunner.HTMLTestRunner(output='..//Reports'))
            # driver.quit()

