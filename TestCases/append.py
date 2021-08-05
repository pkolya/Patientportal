import pyodbc
import openpyxl
from openpyxl.styles import Alignment
import unittest
import HtmlTestRunner
wb = openpyxl.load_workbook('D:/OrthoASC/source/OrthoASC_0140_InsertProcedureOrders_Input.xlsx')
ws = wb.active
cols = ws.max_column
print(ws.cell(617,6).value.replace(" ",""))
string1= """PRE OP: 
1) Shave/clip the operative site if needed.
2) Confirm patient allergies/vitals and follow antibiotic protocol 
3) Follow NSAIDs protocol
4) A DVT risk assessment will be done for patients having general anesthesia greater than 45 minutes. Patients with a score of 2 or greater will have SCDs applied and started in pre-op."""
string3='''PRE OP:
1) Shave/clip the operative site if needed.
2) Confirm patient allergies/vitals and follow antibiotic protocol
3) Follow NSAIDs protocol
4) A DVT risk assessment will be done for patients having general anesthesia greater than 45 minutes. Patients with a score of 2 or greater will have SCDs applied and started in pre-op.'''

string2=string1.replace(" ","")
print (len(string1))
print(len(string2))

string4=string3.replace(" ","")
print (len(string3))
print(len(string4))