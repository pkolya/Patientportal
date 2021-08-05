import pyodbc
import openpyxl
import unittest
import HtmlTestRunner


class TestL(unittest.TestCase):




    # wb=openpyxl.load_workbook('D:/OrthoASC/source/OrthoASC_04_InsertProcedureOrders.xlsx')
    # ws=wb.active
    # cols=ws.max_column
    # ws.cell(1,cols+1).value="procedure_name"
    # ws.cell(1,cols+2).value="actual_sort_order"
    # ws.cell(1,cols+3).value="actual_class"
    # ws.cell(1,cols+4).value="actual_activity_type"
    # ws.cell(1,cols+5).value="actual_ctegory"
    # ws.cell(1,cols+6).value="actualdesc"
    # ws.cell(1,cols+7).value="procname_status"
    # ws.cell(1,cols+8).value="sortorder_status"
    # ws.cell(1,cols+9).value="class_status"
    # ws.cell(1,cols+10).value="activity_type_status"
    # ws.cell(1,cols+11).value="catergory_status"
    # ws.cell(1,cols+12).value="desc_status"
    #specialty="Orthopedic ASC"

        def test_procedureorder(self):
            wb = openpyxl.load_workbook('D:/OrthoASC/source/OrthoASC_0110_InsertProcedure.xlsx')
            ws = wb.worksheets[2]
            cols = ws.max_column

            thisList = []
            for z in range(1, cols + 1):
                ele = ws.cell(1, z).value
                thisList.insert(z, ele)
            newcols = cols + cols
            x = 0
            for y in range(cols + 1, newcols + 1):
                ws.cell(1, y).value = thisList[x]
                x += 1
                b=0
            for status in range(newcols + 1, newcols + cols + 1):
                ws.cell(1, status).value = thisList[b]
                b += 1

            #wb.save('D:/OrthoASC/source/OrthoASC_04_InsertProcedureOrderstest12.xlsx')
            specialty = "Orthopedic ASC"
            query="""select DISTINCT txt_specialty,txt_finding,txt_sort_order from ngkbm_hpi_findings_ where kbm_ind='N' and txt_specialty='Orthopedic ASC' and practice_id='0001'and txt_finding_type='group 1 (mutually exclusive)' order by 3"""
            conn = pyodbc.connect(driver='{SQL Server Native Client 11.0}',
                                               server='NGBLR-03070\MSSQL2014',
                                               database='ACE_Ortho',
                                               uid='sa',pwd='nextgen')
            cursor = conn.cursor()
            cursor.execute(query)
            onerow =cursor.fetchall()
            for x in range(0,len(onerow)):
                data = onerow[x]

           # print("oneline",oneline)
            wb.save('D:/OrthoASC/Results/Results/OrthoASC_04_InsertProcedureInsertCategories_Output.xlsx')

            for x in range(0,len(onerow)):
                elemnt1 = thisList[0]
                target_specialty=onerow[x].txt_specialty
                targetvalue_txt_finding=onerow[x].txt_finding
                #targetvalueclass=onerow[x].class
                target_sort_order = onerow[x].txt_sort_order
                # target_procedure_site = onerow[x].procedure_site
                # target_hpi_reason = onerow[x].hpi_reason
                #targetdesc = onerow[x].description
                ws.cell(row=x+2, column=cols+1).value = target_specialty
                ws.cell(row=x+2, column=cols+2).value=targetvalue_txt_finding
                ws.cell(row=x+2, column=cols+3).value = target_sort_order
                # ws.cell(row=x+2, column=cols+4).value = target_procedure_site
                # ws.cell(row=x+2, column=cols+5).value = target_hpi_reason


                print(onerow[x])
                #print(ws.max_row)
                print(cols)
            wb.save('D:/OrthoASC/Results/Results/OrthoASC_04_InsertProcedureInsertCategories_Output.xlsx')
            for s in range(2,ws.max_row+1):

                if ws.cell(s,1).value==ws.cell(s,cols+1).value:
                    ws.cell(s,cols+cols+1).value="Pass"
                else:
                    ws.cell(s, cols+cols+1).value ="Fail"
                if ws.cell(s,2).value==ws.cell(s,cols+2).value:
                    ws.cell(s,cols+cols+2).value="Pass"
                else:
                    ws.cell(s, cols+cols+2).value ="Fail"
                if ws.cell(s, 3).value == ws.cell(s, cols+3).value:
                    ws.cell(s, cols+cols+3).value = "Pass"
                else:
                    ws.cell(s, cols+cols+3).value = "Fail"
                # if ws.cell(s, 4).value == ws.cell(s, cols+4).value:
                #     ws.cell(s, cols+cols+4).value = "Pass"
                # else:
                #     ws.cell(s, cols+cols+4).value = "Fail"
                # if ws.cell(s, 5).value == ws.cell(s, cols+5).value:
                #     ws.cell(s, cols+cols+5).value = "Pass"
                # else:
                    #ws.cell(s, cols+cols+5).value = "Fail"
            wb.save('D:/OrthoASC/Results/Results/OrthoASC_04_InsertProcedureInsertCategories_Output.xlsx')

        if __name__== '__main__':
            unittest.main(testRunner=HtmlTestRunner.HTMLTestRunner(output='..//Reports'))
         #driver.quit()

