import pyodbc
import openpyxl
import unittest
import HtmlTestRunner


class TestL(unittest.TestCase):


        def test123(self):


            wb=openpyxl.load_workbook('D:/OrthoASC/source/OrthoASC_0170_InsertDocumentationLog_last.xlsx')
            ws = wb.worksheets[1]
            cols = ws.max_column
            specialty = "Orthopedic ASC"

            thisList = []
            for z in range(1, cols + 1):
                ele = ws.cell(1, z).value
                thisList.insert(z, ele)
            newcols = cols + cols
            x = 0
            for y in range(cols + 1, newcols + 1):
                ws.cell(1, y).value = thisList[x]
                x += 1
            b = 0
            for status in range(newcols + 1, newcols + cols + 1):
                ws.cell(1, status).value = thisList[b]
                b += 1

                # specialty = "Orthopedic ASC"
                query = "select entry_type,category,category_order,description,description_order,picklist_type,value_text,value_type,value_order,template_name,value_type_1,value_type_2 from ngkbm_event_log_items_ where specialty='Orthopedic ASC' and practice_id='0001' and kbm_ind='N' order by 1,3,2,5,6,7,9,10"

                conn = pyodbc.connect(driver='{SQL Server Native Client 11.0}',
                                      server='NGBLR-03070\MSSQL2014',
                                      database='ACE_Ortho',
                                      uid='sa', pwd='nextgen')
                cursor = conn.cursor()
                cursor.execute(query)

                onerow = cursor.fetchall()
                for x in range(0, len(onerow)):
                    actual_entry_type = onerow[x].entry_type
                    actual_category = onerow[x].category
                    actual_category_order = onerow[x].category_order
                    actual_description = onerow[x].description
                    actual_description_order = onerow[x].description_order
                    actual_picklist_type= onerow[x].picklist_type
                    actual_value_text = onerow[x].value_text
                    actual_value_type = onerow[x].value_type
                    actual_value_order = onerow[x].value_order
                    actual_template_name = onerow[x].template_name
                    actual_value_type_1 = onerow[x].value_type_1
                    actual_value_type_2 = onerow[x].value_type_2
                    # actual_txt_section_name = onerow[x].txt_section_name
                    # actual_txt_selection_type = onerow[x].txt_selection_type
                    # actual_txt_sentence_id = onerow[x].txt_sentence_id
                    # actual_txt_sub_section_name = onerow[x].txt_sub_section_name
                    # actual_txt_sub_section_id = onerow[x].txt_sub_section_id
                    # actual_txt_sub_selection_number = onerow[x].txt_sub_selection_number
                    # actual_txt_type_parameter = onerow[x].txt_type_parameter

                    ws.cell(row=x + 2, column=cols + 1).value = actual_entry_type
                    ws.cell(row=x + 2, column=cols + 2).value = actual_category
                    ws.cell(row=x + 2, column=cols + 3).value = actual_category_order
                    ws.cell(row=x + 2, column=cols + 4).value = actual_description
                    ws.cell(row=x + 2, column=cols + 5).value = actual_description_order
                    ws.cell(row=x + 2, column=cols + 6).value = actual_picklist_type
                    ws.cell(row=x + 2, column=cols + 7).value = actual_value_text
                    ws.cell(row=x + 2, column=cols + 8).value = actual_value_type
                    ws.cell(row=x + 2, column=cols + 9).value = actual_value_order
                    ws.cell(row=x + 2, column=cols + 10).value = actual_template_name
                    ws.cell(row=x + 2, column=cols + 11).value = actual_value_type_1
                    ws.cell(row=x + 2, column=cols + 12).value = actual_value_type_2
                    # ws.cell(row=x + 2, column=cols + 13).value = actual_txt_section_name
                    # ws.cell(row=x + 2, column=cols + 14).value = actual_txt_selection_type
                    # ws.cell(row=x + 2, column=cols + 15).value = actual_txt_sentence_id
                    # ws.cell(row=x + 2, column=cols + 16).value = actual_txt_sub_section_name
                    # ws.cell(row=x + 2, column=cols + 17).value = actual_txt_sub_section_id
                    # ws.cell(row=x + 2, column=cols + 18).value = actual_txt_sub_selection_number
                    # ws.cell(row=x + 2, column=cols + 19).value = actual_txt_type_parameter
                wb.save('D:/OrthoASC/Results/Results/OrthoASC_02_DocLog_output_log_last3.xlsx')
                colsnew = ws.max_column
                for s in range(2, ws.max_row + 1):
                    for s2 in range(1, cols + 1):
                        if str(ws.cell(s, s2).value) == str(ws.cell(s, cols + s2).value):
                            ws.cell(s, cols + cols + s2).value = "Pass"
                        elif (ws.cell(s, s2).value is None) and (ws.cell(s, cols + s2).value == ""):
                            ws.cell(s, cols + cols + s2).value = "Pass"
                            print("pass", s, s2, ws.cell(s, s2).value)
                        else:
                            ws.cell(s, cols + cols + s2).value = "Fail"
                            print(s, s2, cols + s2, "Expected :", ws.cell(s, s2).value, "Actual :",
                                  ws.cell(s, cols + s2).value)

                wb.save('D:/OrthoASC/Results/Results/OrthoASC_02_DocLog_output_log_last3.xlsx')


if __name__ == '__main__':
    unittest.main(
        testRunner=HtmlTestRunner.HTMLTestRunner(output='D:/Back up/Back up/pyprojects/PATIENTPORTAL/Reports'))
    # driver.quit()

