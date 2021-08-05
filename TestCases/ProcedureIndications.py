import pyodbc
import openpyxl
import unittest
import HtmlTestRunner

class TestL(unittest.TestCase):
    def test123(self):

        wb=openpyxl.load_workbook('D:/OrthoASC/source/OrthoASC_02_InsertIndications.xlsx')
        ws=wb.worksheets[0]
        cols=ws.max_column
        specialty="Orthopedic ASC"

        thisList = []
        for z in range(1, cols + 1):
            ele = ws.cell(1, z).value
            thisList.insert(z, ele)
        newcols = cols + cols
        x = 0
        for y in range(cols + 1, newcols + 1):
            ws.cell(1, y).value = thisList[x]
            x += 1
        b = 0
        for status in range(newcols + 1, newcols + cols + 1):
            ws.cell(1, status).value = thisList[b]
            b += 1


            #specialty = "Orthopedic ASC"
            query="select  txt_specialty,txt_finding_id,txt_sort_order,txt_hpi_reason,txt_caption,txt_finding,txt_sub_selection_id,txt_sub_selection_name,txt_finding_grid1_title,txt_finding_type,txt_group_number,txt_number_of_finding_grids,txt_section_name,txt_selection_type,txt_sentence_id,txt_sub_section_name,txt_sub_section_id,txt_sub_selection_number,txt_type_parameter from ngkbm_hpi_findings_ where txt_specialty ='Orthopedic ASC' AND practice_id='0001' AND txt_finding_type in ('multichoice single','multichoice multiple') order by 2,3 "
            conn = pyodbc.connect(driver='{SQL Server Native Client 11.0}',
                                               server='NGBLR-03070\MSSQL2014',
                                               database='ACE_Ortho',
                                               uid='sa',pwd='nextgen')
            cursor = conn.cursor()
            cursor.execute(query)

            onerow=cursor.fetchall()
            for x in range(0,len(onerow)):
                actual_Specialty=onerow[x].txt_specialty
                actual_txt_finding_id=onerow[x].txt_finding_id
                actual_txt_sort_order=onerow[x].txt_sort_order
                actual_txt_hpi_reason = onerow[x].txt_hpi_reason
                actual_txt_caption = onerow[x].txt_caption
                actual_txt_finding = onerow[x].txt_finding
                actual_txt_sub_selection_id=onerow[x].txt_sub_selection_id
                actual_txt_sub_selection_name = onerow[x].txt_sub_selection_name
                actual_txt_finding_grid1_title = onerow[x].txt_finding_grid1_title
                actual_txt_finding_type = onerow[x].txt_finding_type
                actual_txt_group_number = onerow[x].txt_group_number
                actual_txt_number_of_findings_grids = onerow[x].txt_number_of_finding_grids
                actual_txt_section_name = onerow[x].txt_section_name
                actual_txt_selection_type = onerow[x].txt_selection_type
                actual_txt_sentence_id = onerow[x].txt_sentence_id
                actual_txt_sub_section_name = onerow[x].txt_sub_section_name
                actual_txt_sub_section_id = onerow[x].txt_sub_section_id
                actual_txt_sub_selection_number = onerow[x].txt_sub_selection_number
                actual_txt_type_parameter = onerow[x].txt_type_parameter
                
                ws.cell(row=x+2, column=cols+1).value=actual_Specialty
                ws.cell(row=x+2, column=cols+2).value = actual_txt_finding_id
                ws.cell(row=x+2, column=cols+3).value = actual_txt_sort_order
                ws.cell(row=x+2, column=cols+4).value = actual_txt_hpi_reason
                ws.cell(row=x+2, column=cols+5).value = actual_txt_caption
                ws.cell(row=x+2, column=cols+6).value = actual_txt_finding
                ws.cell(row=x+2, column=cols+7).value = actual_txt_sub_selection_id
                ws.cell(row=x+2, column=cols+8).value = actual_txt_sub_selection_name
                ws.cell(row=x+2, column=cols+9).value = actual_txt_finding_grid1_title
                ws.cell(row=x+2, column=cols+10).value = actual_txt_finding_type
                ws.cell(row=x+2, column=cols+11).value = actual_txt_group_number
                ws.cell(row=x+2, column=cols+12).value = actual_txt_number_of_findings_grids
                ws.cell(row=x+2, column=cols+13).value = actual_txt_section_name
                ws.cell(row=x+2, column=cols+14).value = actual_txt_selection_type
                ws.cell(row=x+2, column=cols+15).value = actual_txt_sentence_id
                ws.cell(row=x+2, column=cols+16).value = actual_txt_sub_section_name
                ws.cell(row=x+2, column=cols+17).value = actual_txt_sub_section_id
                ws.cell(row=x+2, column=cols+18).value = actual_txt_sub_selection_number
                ws.cell(row=x+2, column=cols+19).value = actual_txt_type_parameter
            wb.save('D:/OrthoASC/Results/Results/OrthoASC_02_InsertIndicationsz_output.xlsx')
            colsnew = ws.max_column
            for s in range(2, ws.max_row + 1):
                for s2 in range(1,cols+1):
                    if ws.cell(s, s2).value == ws.cell(s, cols + s2).value:
                        ws.cell(s, cols + cols + s2).value = "Pass"
                    elif (ws.cell(s, s2).value== None) and (ws.cell(s, cols + s2).value==""):
                         ws.cell(s, cols + cols + s2).value = "Pass"
                         print("pass",s,s2,ws.cell(s, s2).value)
                    else:
                        ws.cell(s, cols + cols + s2).value = "Fail"
                        print(s,s2,cols + s2,"Expected :",ws.cell(s, s2).value,"Actual :", ws.cell(s, cols+s2).value)


            wb.save('D:/OrthoASC/Results/Results/OrthoASC_02_InsertIndicationsz_output.xlsx')


if __name__== '__main__':
    unittest.main(testRunner=HtmlTestRunner.HTMLTestRunner(output='D:/Back up/Back up/pyprojects/PATIENTPORTAL/Reports'))
        # driver.quit()

