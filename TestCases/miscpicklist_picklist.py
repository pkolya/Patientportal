import pyodbc
import openpyxl
import unittest
import HtmlTestRunner


class TestL(unittest.TestCase):


        def test123(self):


            wb = openpyxl.load_workbook('D:/OrthoASC/source/OrthoASC_0190_InsertMiscPicklists.xlsx')
            ws = wb.worksheets[0]
            cols = ws.max_column
            #specialty = "Orthopedic ASC"

            thisList = []
            for z in range(1, cols + 1):
                ele = ws.cell(1, z).value
                thisList.insert(z, ele)
            newcols = cols + cols
            x = 0
            for y in range(cols + 1, newcols + 1):
                ws.cell(1, y).value = thisList[x]
                x += 1
            b = 0
            for status in range(newcols + 1, newcols + cols + 1):
                ws.cell(1, status).value = thisList[b]
                b += 1

                # specialty = "Orthopedic ASC"
                query = "select mstr_list_type,mstr_list_item_desc,create_timestamp from mstr_lists where mstr_list_type in ('Flow rate','postop_mangmnt') and create_timestamp like '%2021%'  order by 2"

                conn = pyodbc.connect(driver='{SQL Server Native Client 11.0}',
                                      server='NGBLR-03070\MSSQL2014',
                                      database='ACE_Ortho',
                                      uid='sa', pwd='nextgen')
                cursor = conn.cursor()
                cursor.execute(query)

                onerow = cursor.fetchall()
                for x in range(0, len(onerow)):
                    mstr_list_type = onerow[x].mstr_list_type
                    mstr_list_item_desc = onerow[x].mstr_list_item_desc
                    #description = onerow[x].description

                    ws.cell(row=x + 2, column=cols + 1).value = mstr_list_type
                    ws.cell(row=x + 2, column=cols + 2).value = mstr_list_item_desc
                    #ws.cell(row=x + 2, column=cols + 3).value = description
                wb.save('D:/OrthoASC/Results/Results/OrthoASC_0190_InsertMiscPicklists_Results.xlsx')
                colsnew = ws.max_column
                for s in range(2, ws.max_row + 1):
                    for s2 in range(1, cols + 1):
                        if str(ws.cell(s, s2).value) == str(ws.cell(s, cols + s2).value):
                            ws.cell(s, cols + cols + s2).value = "Pass"
                        elif (ws.cell(s, s2).value is None) and (ws.cell(s, cols + s2).value == ""):
                            ws.cell(s, cols + cols + s2).value = "Pass"
                            print("pass", s, s2, ws.cell(s, s2).value)
                        else:
                            ws.cell(s, cols + cols + s2).value = "Fail"
                            print(s, s2, cols + s2, "Expected :", ws.cell(s, s2).value, "Actual :",
                                  ws.cell(s, cols + s2).value)

                wb.save('D:/OrthoASC/Results/Results/OrthoASC_0190_InsertMiscPicklists_Results.xlsx')


if __name__ == '__main__':
    unittest.main(
        testRunner=HtmlTestRunner.HTMLTestRunner(output='D:/Back up/Back up/pyprojects/PATIENTPORTAL/Reports'))
    # driver.quit()

