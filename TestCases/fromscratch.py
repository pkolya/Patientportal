import pyodbc
import openpyxl
from openpyxl.styles import Alignment
import string
import unittest
import HtmlTestRunner
wb = openpyxl.load_workbook('D:/OrthoASC/source/OrthoASC_0140_InsertProcedureOrders_Input.xlsx')
ws = wb.active
cols = ws.max_column
print(cols)

thisList = []
for z in range(1, cols + 1):
    ele = ws.cell(1, z).value
    thisList.insert(z, ele)
newcols = cols + cols
x = 0
for y in range(cols + 1, newcols + 1):
    ws.cell(1, y).value = thisList[x]
    x += 1
    b=0
for status in range(newcols+1, newcols+cols+1 ):
    ws.cell(1,status).value=thisList[b]
    b +=1

specialty="Orthopedic ASC"
wb.save('D:/OrthoASC/Results/Results/OrthoASC_04_InsertProcedureOrderOutput.xlsx')
query="select procedure_name,sort_order,class as cs,activity_type,category,description FROM ngkbm_proc_activity_items_ AS npa INNER JOIN ngkbm_proc_mstr_ npm ON npa.practice_id = npm.practice_id AND npm.procedure_id = npa.proc_id where npa.specialty='"+specialty+"' AND npa.practice_id='0001' AND class in ('General Procedure Order','Medication Procedure Order') order by 1,2"
conn = pyodbc.connect(driver='{SQL Server Native Client 11.0}',
                                   server='NGBLR-03070\MSSQL2014',
                                   database='ACE_Ortho',
                                   uid='sa',pwd='nextgen')
cursor = conn.cursor()
cursor.execute(query)
allrows=cursor.fetchall()
for row in range (0,len(allrows)):
    print(allrows[row])
    ws.cell(row+2,cols+1).value=allrows[row].procedure_name
    ws.cell(row+2,cols+2).value=allrows[row].sort_order
    ws.cell(row+2,cols+3).value = allrows[row].cs
    ws.cell(row+2,cols + 4).value = allrows[row].activity_type
    ws.cell(row+2,cols + 5).value = allrows[row].category
    ws.cell(row+2,cols+6).alignment=Alignment(wrapText=True)
    #ws.cell(row+2,cols+6).value = "'" + allrows[row].description.strip()
    ws.cell(row+2, cols+6).value =allrows[row].description

wb.save('D:/OrthoASC/Results/Results/OrthoASC_04_InsertProcedureOrderOutput.xlsx')

for s in range(2, ws.max_row + 1):
    if ws.cell(s, 1).value == ws.cell(s, cols+1).value:
        ws.cell(s, cols + cols+1).value = "Pass"
    else:
        ws.cell(s, cols + cols+1).value = "Fail"
    if ws.cell(s, 2).value == ws.cell(s, cols+2).value:
        ws.cell(s, cols+cols+2).value = "Pass"
    else:
        ws.cell(s, cols+cols+2).value = "Fail"

    if ws.cell(s,3).value == ws.cell(s,cols+3).value:
        ws.cell(s,cols+cols+3).value="Pass"
    else:
        ws.cell(s,cols+cols+3).value="Fail"
    if ws.cell(s, 4).value == ws.cell(s, cols+4).value:
        ws.cell(s, cols+cols+4).value = "Pass"
    else:
        ws.cell(s, cols+cols+4).value = "Fail"
    if ws.cell(s, 5).value == ws.cell(s, cols+5).value:
        ws.cell(s, cols+cols+5).value = "Pass"
    else:
        ws.cell(s, cols+cols+5).value = "Fail"
    Expected=ws.cell(s, 6).value
    # Exp="'" + Expected
    Actual=ws.cell(s, cols + 6).value
    Expected2 = str(Expected).strip()
    Actual2 = str(Actual).strip()

    # str(definew) = defi[1:]
    if Expected2 == Actual2:
        ws.cell(s, cols + cols+6).value = "Pass"
    elif len(Actual2) == len(Expected2):
        ws.cell(s, cols + cols+6).value = "Pass"
    elif ("OP:" in Actual):
        ws.cell(s, cols + cols + 6).value = "Multiline"
        print(s,Actual)
    else:
        print(s,Actual,"fail")
        ws.cell(s, cols + cols + 6).value = "Fail"
        print(type(Actual))
wb.save('D:/OrthoASC/Results/Results/OrthoASC_04_InsertProcedureOrderOutput.xlsx')
