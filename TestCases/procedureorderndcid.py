import pyodbc
import openpyxl
from openpyxl.styles import Alignment
import unittest
import HtmlTestRunner
wb = openpyxl.load_workbook('D:/OrthoASC/source/OrthoASC_0140_InsertProcedureOrders_last.xlsx')
ws=wb.worksheets[1]
# ws = wb.get_sheet_by_name("")
cols = ws.max_column
print(cols)

thisList = []
for z in range(1, cols + 1):
    ele = ws.cell(1, z).value
    thisList.insert(z, ele)
newcols = cols + cols
x = 0
for y in range(cols + 1, newcols + 1):
    ws.cell(1, y).value = thisList[x]
    x += 1
    b=0
for status in range(newcols+1, newcols+cols+1 ):
    ws.cell(1,status).value=thisList[b]
    b +=1

    specialty="Orthopedic ASC"
wb.save('D:/OrthoASC/Results/Results/OrthoASC_0140_InsertProcedureOrders_NDCID_output_last.xlsx')
query="SELECT DISTINCT proc_act.description,proc_act.dose, proc_act.entry_type,proc_act.medication,fm.generic_name,proc_act.ndc_id,proc_act.route,proc_act.unit FROM ngkbm_proc_activity_items_ AS proc_act INNER JOIN dbo.ngkbm_proc_mstr_ AS proc_mstr ON (proc_mstr.procedure_id = proc_act.proc_id) LEFT OUTER JOIN fdb_medication AS fm ON proc_act.ndc_id = fm.ndc_id WHERE proc_act.class = 'Medication Procedure Order' AND proc_act.practice_id = '0001' AND proc_act.specialty ='"+specialty+"' AND Proc_act.ndc_id is not NULL and proc_act.kbm_ind='N' order by 4,1,2"
conn = pyodbc.connect(driver='{SQL Server Native Client 11.0}',
                                   server='NGBLR-03070\MSSQL2014',
                                   database='ACE_Ortho',
                                   uid='sa',pwd='nextgen')
cursor = conn.cursor()
cursor.execute(query)
allrows=cursor.fetchall()
for row in range (0,len(allrows)):
    print(row)
    ws.cell(row + 2, cols + 1).alignment = Alignment(wrapText=True)
    ws.cell(row+2,cols+1).value=allrows[row].description
    ws.cell(row + 2, cols + 2).alignment=Alignment(wrapText=True)
    ws.cell(row+2,cols+2).value=allrows[row].dose
    ws.cell(row+2,cols+3).value = allrows[row].entry_type
    ws.cell(row+2,cols + 4).value = allrows[row].medication
    ws.cell(row+2,cols + 5).value = allrows[row].ndc_id
    ws.cell(row+2,cols + 6).value = allrows[row].route
    ws.cell(row+2,cols + 7).value =allrows[row].unit
wb.save('D:/OrthoASC/Results/Results/OrthoASC_0140_InsertProcedureOrders_NDCID_output_last.xlsx')

for s in range(2, ws.max_row + 1):
    if ws.cell(s, 1).value == ws.cell(s, cols + 1).value:
        ws.cell(s, cols + cols+1).value = "Pass"
    else:
        ws.cell(s, cols +cols+1).value = "Fail"
    if str(ws.cell(s, 2).value) == str(ws.cell(s, cols + 2).value):
        ws.cell(s, cols+cols+2).value = "Pass"
    else:
        ws.cell(s, cols+cols+2).value = "Fail"
        print(ws.cell(s, 2).value)
        print(ws.cell(s, cols + 2).value)

    if ws.cell(s,3).value == ws.cell(s,cols+3).value:
        ws.cell(s,cols+cols+3).value="Pass"
    else:
        ws.cell(s,cols+cols+3).value="Fail"
    if ws.cell(s, 4).value == ws.cell(s, cols + 4).value:
        ws.cell(s, cols+cols+4).value = "Pass"
    else:
        ws.cell(s, cols+cols+4).value = "Fail"
    if ws.cell(s, 5).value == ws.cell(s, cols + 5).value:
        ws.cell(s, cols+cols+5).value = "Pass"
    else:
        ws.cell(s, cols+cols+5).value = "Fail"
    # Expected=ws.cell(s, 6).value
    # Exp="'" + Expected
    # Actual=ws.cell(s, cols + 6).value

    # str(definew) = defi[1:]
    if ws.cell(s,6).value == ws.cell(s,cols+6).value:
        ws.cell(s, cols +cols+6).value = "Pass"
    else:
        ws.cell(s, cols +cols+6).value = "Fail"
    if ws.cell(s, 7).value == ws.cell(s, cols + 7).value:
        ws.cell(s, cols + cols + 7).value = "Pass"
    else:
        ws.cell(s, cols + cols + 7).value = "Fail"

wb.save('D:/OrthoASC/Results/Results/OrthoASC_0140_InsertProcedureOrders_NDCID_output_last.xlsx')
