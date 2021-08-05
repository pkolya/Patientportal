import pyodbc
import openpyxl

wb=openpyxl.load_workbook('D:/OrthoASC/source/OrthoASC_04_InsertProcedureOrders.xlsx')
ws=wb.active
cols=ws.max_column
ws.cell(1,cols+1).value="specialty"
ws.cell(1,cols+2).value="actual_panel_name"
ws.cell(1,cols+3).value="actual_sort_order"
ws.cell(1,cols+4).value="timeout_status"
ws.cell(1,cols+5).value="panel_status"
ws.cell(1,cols+6).value="sort_order_status"
specialty="Orthopedic ASC"

query="select specialty,panel_name,sort_order,timeout_element FROM ngkbm_proc_timeout_items_ where practice_id='0001'AND specialty ='"+specialty+"' order by 1"
conn = pyodbc.connect(driver='{SQL Server Native Client 11.0}',
                                   server='NGBLR-03070\MSSQL2014',
                                   database='ACE_Ortho',
                                   uid='sa',pwd='nextgen')
cursor = conn.cursor()
cursor.execute(query)

onerow=cursor.fetchall()
for x in range(0,len(onerow)):
    targetvaluetimeout=onerow[x].timeout_element
    targetvaluepanelname=onerow[x].panel_name
    targetsortorder = onerow[x].sort_order
    ws.cell(row=x+2, column=cols+1).value=targetvaluetimeout
    ws.cell(row=x+2, column=cols+2).value = targetvaluepanelname
    ws.cell(row=x+2, column=cols+3).value = targetsortorder
    print(onerow[x])
    print(ws.max_row)
wb.save('D:/OrthoASC/source/OrthoASC_04_InsertProcedureOrders.xlsx')
for s in range(2,ws.max_row+1):
    if ws.cell(s,4).value==ws.cell(s,cols+1).value:
        ws.cell(s,cols+4).value="Pass"
    else:
        ws.cell(s, cols+4).value ="Fail"
    if ws.cell(s,2).value==ws.cell(s,cols+2).value:
        ws.cell(s,cols+5).value="Pass"
    else:
        ws.cell(s, cols+5).value ="Fail"
    if ws.cell(s, 3).value == ws.cell(s, cols+3).value:
        ws.cell(s, cols+6).value = "Pass"
    else:
        ws.cell(s, cols+6).value = "Fail"
wb.save('D:/OrthoASC/source/OrthoASC_04_InsertProcedureOrders.xlsx')
