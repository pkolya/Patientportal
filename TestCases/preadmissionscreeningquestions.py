import pyodbc
import openpyxl
import unittest
import HtmlTestRunner


class TestL(unittest.TestCase):
    def test123(self):

        wb = openpyxl.load_workbook('D:/OrthoASC/source/OrthoASC_0180_InsertPreadmission.xlsx')
        ws = wb.worksheets[1]
        cols = ws.max_column
        specialty = "Orthopedic ASC"

        thisList = []
        for z in range(1, cols + 1):
            ele = ws.cell(1, z).value
            thisList.insert(z, ele)
        newcols = cols + cols
        x = 0
        for y in range(cols + 1, newcols + 1):
            ws.cell(1, y).value = thisList[x]
            x += 1
        b = 0
        for status in range(newcols + 1, newcols + cols + 1):
            ws.cell(1, status).value = thisList[b]
            b += 1

            # specialty = "Orthopedic ASC"
            query = """SELECT m.procedure_name, q.chk_schlde_review, q.schlde_section, q.schlde_questions, q.schlde_answer, q.schlde_ans_type, q.schlde_val_type_size,
q.section_order, q.question_order, q.sort_order, q.schlde_primary_section, q.schlde_primary_question, q.schlde_primary_answer, q.schlde_alert,
i.category AS act_category, i.description AS act_description, i.description_order AS act_description_order, i.section_order AS act_section_order, i.sort_order AS act_sort_order,
i2.category AS instruct_category, i2.description AS instruct_description
FROM ngkbm_screening_questions_ q
JOIN ngkbm_screening_proc_xref_ x on q.screening_question_id = x.screening_question_id AND q.practice_id = x.practice_id
JOIn ngkbm_proc_mstr_ m on x.practice_id = m.practice_id AND x.procedure_id = m.procedure_id
LEFT JOIN ngkbm_proc_activity_items_ i ON q.practice_id = i.practice_id AND q.chart_check_id = i.instruction_id AND i.activity_type = 'Chart Check'
LEFT JOIN ngkbm_proc_activity_items_ i2 ON q.practice_id = i2.practice_id AND q.instruction_id = i2.instruction_id AND i2.activity_type = 'Patient Instructions'
WHERE q.kbm_ind <> 'D'
AND q.practice_id = '0001'
AND q.schlde_specality = 'orthopedic asc' ORDER BY m.procedure_name, q.section_order, q.question_order, q.sort_order"""
            conn = pyodbc.connect(driver='{SQL Server Native Client 11.0}',
                                  server='NGBLR-03070\MSSQL2014',
                                  database='ACE_Ortho',
                                  uid='sa', pwd='nextgen')
            cursor = conn.cursor()
            cursor.execute(query)

            onerow = cursor.fetchall()
            for x in range(0, len(onerow)):
                procedure_name = onerow[x].procedure_name
                chk_schlde_review = onerow[x].chk_schlde_review
                schlde_section = onerow[x].schlde_section
                schlde_questions = onerow[x].schlde_questions
                schlde_answer = onerow[x].schlde_answer
                schlde_ans_type = onerow[x].schlde_ans_type
                schlde_val_type_size = onerow[x].schlde_val_type_size
                section_order = onerow[x].section_order
                question_order = onerow[x].question_order
                sort_order = onerow[x].sort_order
                schlde_primary_section = onerow[x].schlde_primary_section
                schlde_primary_question = onerow[x].schlde_primary_question
                schlde_primary_answer = onerow[x].schlde_primary_answer
                schlde_alert = onerow[x].schlde_alert
                act_category = onerow[x].act_category
                act_description = onerow[x].act_description
                act_description_order = onerow[x].act_description_order
                act_section_order = onerow[x].act_section_order
                act_sort_order = onerow[x].act_sort_order
                instruct_category = onerow[x].instruct_category
                instruct_description = onerow[x].instruct_description

                ws.cell(row=x + 2, column=cols + 1).value = procedure_name
                ws.cell(row=x + 2, column=cols + 2).value = chk_schlde_review
                ws.cell(row=x + 2, column=cols + 3).value = schlde_section
                ws.cell(row=x + 2, column=cols + 4).value = schlde_questions
                ws.cell(row=x + 2, column=cols + 5).value = schlde_answer
                ws.cell(row=x + 2, column=cols + 6).value = schlde_ans_type
                ws.cell(row=x + 2, column=cols + 7).value = schlde_val_type_size
                ws.cell(row=x + 2, column=cols + 8).value = section_order
                ws.cell(row=x + 2, column=cols + 9).value = question_order
                ws.cell(row=x + 2, column=cols + 10).value = sort_order
                ws.cell(row=x + 2, column=cols + 11).value = schlde_primary_section
                ws.cell(row=x + 2, column=cols + 12).value = schlde_primary_question
                ws.cell(row=x + 2, column=cols + 13).value = schlde_primary_answer
                ws.cell(row=x + 2, column=cols + 14).value = schlde_alert
                ws.cell(row=x + 2, column=cols + 15).value = act_category
                ws.cell(row=x + 2, column=cols + 16).value = act_description
                ws.cell(row=x + 2, column=cols + 17).value = act_description_order
                ws.cell(row=x + 2, column=cols + 18).value = act_section_order
                ws.cell(row=x + 2, column=cols + 19).value = act_sort_order
                ws.cell(row=x + 2, column=cols + 20).value = instruct_category
                ws.cell(row=x + 2, column=cols + 21).value = instruct_description


            wb.save('D:/OrthoASC/Results/Results/OrthoASC_0180_InsertPreadmission_last.xlsx')
            colsnew = ws.max_column
            for s in range(2, ws.max_row + 1):
                for s2 in range(1, cols + 1):
                    if ws.cell(s, s2).value == ws.cell(s, cols + s2).value:
                        ws.cell(s, cols + cols + s2).value = "Pass"
                    elif (ws.cell(s, s2).value == None) and (ws.cell(s, cols + s2).value == ""):
                        ws.cell(s, cols + cols + s2).value = "Pass"
                        #print("pass", s, s2, ws.cell(s, s2).value)
                    elif (ws.cell(s, s2).value =="NULL") and (ws.cell(s, cols + s2).value is None):
                        ws.cell(s, cols + cols + s2).value = "Pass"
                    else:
                        ws.cell(s, cols + cols + s2).value = "Fail"
                        #print(s, s2, cols + s2, "Expected :", ws.cell(s, s2).value, "Actual :",
                              #ws.cell(s, cols + s2).value)

            wb.save('D:/OrthoASC/Results/Results/OrthoASC_0180_InsertPreadmission_last.xlsx')


if __name__ == '__main__':
    unittest.main(
        testRunner=HtmlTestRunner.HTMLTestRunner(output='D:/Back up/Back up/pyprojects/PATIENTPORTAL/Reports'))
    # driver.quit()

